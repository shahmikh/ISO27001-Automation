# src/exporter.py

import json
from pathlib import Path
import pandas as pd
from datetime import datetime


# -------------------------
# Loaders
# -------------------------

def load_results(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)

def load_mappings(path):
    return pd.read_csv(path)

def load_gaps(path):
    return pd.read_csv(path)


# -------------------------
# Summary Builder
# -------------------------

def build_summary(results):
    total = len(results)
    compliant = sum(1 for r in results if r["status"] == "Compliant")
    partial = sum(1 for r in results if r["status"] == "Partially Compliant")
    not_compliant = sum(1 for r in results if r["status"] == "Not Compliant")

    compliance_pct = round((compliant / total) * 100, 2) if total else 0.0

    total_weight = sum(r["weight"] for r in results)
    achieved_weight = sum(r["weighted_score"] for r in results)

    weighted_compliance = (
        round((achieved_weight / total_weight) * 100, 2)
        if total_weight else 0.0
    )

    return {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "total_controls": total,
        "compliant": compliant,
        "partially_compliant": partial,
        "not_compliant": not_compliant,
        "compliance_pct": compliance_pct,
        "weighted_compliance": weighted_compliance
    }


# -------------------------
# Excel Exporter
# -------------------------

def export_excel(results, mappings_df, gaps_df, out_path):
    """
    Create Excel dashboard with:
      - Summary sheet
      - Controls sheet
      - Mappings sheet
      - Gaps sheet
      - Charts sheet with bar charts
    """

    import openpyxl
    from openpyxl.chart import BarChart, Reference

    summary = build_summary(results)

    results_df = pd.DataFrame(results)

    # ---- Step 1: Create workbook ----
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame([summary]).to_excel(writer, sheet_name="Summary", index=False)
        results_df.to_excel(writer, sheet_name="Controls", index=False)
        mappings_df.to_excel(writer, sheet_name="Mappings", index=False)
        gaps_df.to_excel(writer, sheet_name="Gaps", index=False)

    # ---- Step 2: Add charts ----
    wb = openpyxl.load_workbook(out_path)

    if "Charts" not in wb.sheetnames:
        wb.create_sheet("Charts")
    ws = wb["Charts"]

    # -------------------------------
    # Chart 1: Compliance Count
    # -------------------------------
    ws["A1"] = "Status"
    ws["B1"] = "Count"

    ws["A2"] = "Compliant"
    ws["B2"] = summary["compliant"]

    ws["A3"] = "Partially Compliant"
    ws["B3"] = summary["partially_compliant"]

    ws["A4"] = "Not Compliant"
    ws["B4"] = summary["not_compliant"]

    chart1 = BarChart()
    chart1.title = "Compliance Status Distribution"
    chart1.x_axis.title = "Status"
    chart1.y_axis.title = "Count"

    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    labels = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(labels)
    chart1.width = 18
    chart1.height = 10

    ws.add_chart(chart1, "D2")

    # -------------------------------
    # Chart 2: Weighted Compliance %
    # -------------------------------
    ws["A7"] = "Weighted Compliance %"
    ws["B7"] = summary["weighted_compliance"]

    chart2 = BarChart()
    chart2.title = "Weighted Compliance Score"
    chart2.x_axis.title = "Metric"
    chart2.y_axis.title = "Percentage"

    data2 = Reference(ws, min_col=2, min_row=7, max_row=7)
    labels2 = Reference(ws, min_col=1, min_row=7, max_row=7)

    chart2.add_data(data2)
    chart2.set_categories(labels2)
    chart2.width = 12
    chart2.height = 8

    ws.add_chart(chart2, "D20")

    wb.save(out_path)

    return summary


# -------------------------
# PDF Exporter
# -------------------------

def export_pdf(summary, out_path):
    """
    Professional one-page PDF summary using reportlab.
    Includes header, summary table, and footer.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(out_path), pagesize=letter)
    width, height = letter

    left = 50
    top = height - 50
    line_height = 18

    # ----- Header -----
    c.setFont("Helvetica-Bold", 20)
    c.drawString(left, top, "ISO 27001 Compliance Summary")

    c.setFont("Helvetica", 11)
    c.drawString(left, top - 25, f"Generated at (UTC): {summary['generated_at']}")

    # ----- Summary Table -----
    c.setFont("Helvetica-Bold", 12)
    c.drawString(left, top - 60, "Compliance Metrics:")

    table_data = [
        ["Metric", "Value"],
        ["Total Controls", summary["total_controls"]],
        ["Compliant", summary["compliant"]],
        ["Partially Compliant", summary["partially_compliant"]],
        ["Not Compliant", summary["not_compliant"]],
        ["Compliance %", f"{summary['compliance_pct']}%"],
        ["Weighted Compliance %", f"{summary['weighted_compliance']}%"],
    ]

    # Draw table
    c.setFont("Helvetica", 10)
    y = top - 80

    for row in table_data:
        c.drawString(left, y, str(row[0]))
        c.drawRightString(left + 300, y, str(row[1]))
        y -= line_height

    # ----- Footer -----
    c.setFont("Helvetica-Oblique", 9)
    c.setFillColor(colors.grey)
    c.drawString(left, 60, "Generated by: ISO 27001 Compliance Automation Engine (Python)")
    c.drawString(left, 45, "This document provides a high-level compliance overview.")
    c.setFillColor(colors.black)

    c.showPage()
    c.save()


# -------------------------
# Main Entry
# -------------------------

def main():
    project_root = Path(__file__).resolve().parents[1]

    results_path = project_root / "data" / "results.json"
    mappings_path = project_root / "data" / "mappings.csv"
    gaps_path = project_root / "data" / "gaps.csv"

    out_excel = project_root / "data" / "report.xlsx"
    out_pdf = project_root / "data" / "summary.pdf"

    print("📦 Loading data...")
    results = load_results(results_path)
    mappings_df = load_mappings(mappings_path)
    gaps_df = load_gaps(gaps_path)

    print("📊 Exporting Excel report...")
    summary = export_excel(results, mappings_df, gaps_df, out_excel)
    print(f"✔ Excel saved: {out_excel}")

    print("📝 Exporting PDF summary...")
    export_pdf(summary, out_pdf)
    print(f"✔ PDF saved:   {out_pdf}")

    print("\n🎉 Export complete!")
    print(f"- Excel Dashboard: {out_excel}")
    print(f"- PDF Summary:     {out_pdf}")


if __name__ == "__main__":
    main()

